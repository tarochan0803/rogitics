#!/usr/bin/env python
"""Import site-point teacher data from an Excel workbook.

The expected workbook is intentionally simple:
  truck result / latitude / longitude

The importer treats recorded truck classes as observed positives. Larger
unrecorded classes are emitted only as weak negatives; they should not be used
as hard NG labels until the business process confirms that they were not sent
because they could not enter.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import tempfile
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_OUT = ROOT / "runtime" / "teacher_data" / "teacher_points.json"
DEFAULT_CSV_OUT = ROOT / "runtime" / "teacher_data" / "teacher_points.csv"
TRUCK_CLASSES = [2, 3, 4, 10]
PRESET_BY_CLASS = {
    2: "2t_flat",
    3: "3t_flat",
    4: "4t_flat",
    10: "10t_unic",
}


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import teacher site points from xlsx.")
    parser.add_argument("--xlsx", default=None, help="Input xlsx. Default: first non-lock *.xlsx in repo root.")
    parser.add_argument("--sheet", default=None, help="Worksheet name. Default: first sheet.")
    parser.add_argument("--out", default=str(DEFAULT_OUT), help="Output JSON path.")
    parser.add_argument("--csv-out", default=str(DEFAULT_CSV_OUT), help="Output CSV summary path.")
    parser.add_argument("--dry-run", action="store_true", help="Validate and print summary without writing.")
    parser.add_argument("--selfcheck", action="store_true", help="Run importer selfcheck.")
    return parser.parse_args(argv)


def find_default_xlsx() -> Path:
    files = [p for p in ROOT.glob("*.xlsx") if not p.name.startswith("~$")]
    if not files:
      raise FileNotFoundError("No *.xlsx workbook found in repo root.")
    preferred = [p for p in files if "データ" in p.name or "teacher" in p.name.lower()]
    return preferred[0] if preferred else files[0]


def normalize_header(value: object) -> str:
    return str(value or "").strip().lower()


def find_column(headers: list[object], candidates: list[str], fallback: int) -> int:
    normalized = [normalize_header(h) for h in headers]
    for i, value in enumerate(normalized):
        for cand in candidates:
            if cand in value:
                return i
    return fallback


def parse_truck_class(value: object) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and int(value) == float(value):
        n = int(value)
        return n if n in PRESET_BY_CLASS else None
    text = str(value).strip().lower()
    match = re.search(r"(10|[234])\s*t?", text)
    if not match:
        return None
    n = int(match.group(1))
    return n if n in PRESET_BY_CLASS else None


def read_workbook_rows(xlsx: Path, sheet: str | None = None) -> tuple[str, list[dict]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is required. Use the project venv or install openpyxl.") from exc

    workbook = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    try:
        worksheet = workbook[sheet] if sheet else workbook.worksheets[0]
        rows_iter = worksheet.iter_rows(values_only=True)
        headers = list(next(rows_iter, []))
        truck_col = find_column(headers, ["トラック", "truck", "車格", "結果"], 0)
        lat_col = find_column(headers, ["緯度", "lat"], 1)
        lng_col = find_column(headers, ["経度", "lng", "lon"], 2)

        rows: list[dict] = []
        for row_no, row in enumerate(rows_iter, start=2):
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue
            def cell(index: int) -> object:
                return row[index] if index < len(row) else None

            truck_class = parse_truck_class(cell(truck_col))
            try:
                lat = float(cell(lat_col))
                lng = float(cell(lng_col))
            except (TypeError, ValueError):
                rows.append({"row": row_no, "error": "lat/lng is not numeric", "raw": list(row)})
                continue
            if truck_class is None:
                rows.append({"row": row_no, "error": "truck class is not 2/3/4/10", "raw": list(row)})
                continue
            if not (-90 <= lat <= 90 and -180 <= lng <= 180):
                rows.append({"row": row_no, "error": "lat/lng out of range", "raw": list(row)})
                continue
            rows.append({"row": row_no, "truckClass": truck_class, "lat": lat, "lng": lng})
        return worksheet.title, rows
    finally:
        workbook.close()


def build_points(rows: list[dict]) -> tuple[list[dict], list[dict]]:
    errors = [r for r in rows if "error" in r]
    valid = [r for r in rows if "error" not in r]
    groups: dict[tuple[float, float], list[dict]] = defaultdict(list)
    for row in valid:
        groups[(round(float(row["lat"]), 7), round(float(row["lng"]), 7))].append(row)

    points: list[dict] = []
    for idx, ((lat, lng), members) in enumerate(sorted(groups.items(), key=lambda item: min(r["row"] for r in item[1])), start=1):
        observed = sorted({int(m["truckClass"]) for m in members})
        max_observed = max(observed)
        inferred_pass = [n for n in TRUCK_CLASSES if n <= max_observed]
        weak_negative = [n for n in TRUCK_CLASSES if n > max_observed]
        source_rows = [int(m["row"]) for m in members]
        points.append({
            "id": f"teacher-site-{idx:04d}",
            "name": f"teacher site {idx:04d}",
            "lat": lat,
            "lng": lng,
            "observedTruckClasses": observed,
            "observedPositivePresets": [PRESET_BY_CLASS[n] for n in observed],
            "inferredPassablePresets": [PRESET_BY_CLASS[n] for n in inferred_pass],
            "weakNegativePresets": [PRESET_BY_CLASS[n] for n in weak_negative],
            "sourceRows": source_rows,
            "duplicateCount": len(members),
        })
    return points, errors


def write_outputs(payload: dict, out_path: Path, csv_path: Path) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow([
            "id", "lat", "lng", "observedTruckClasses", "observedPositivePresets",
            "inferredPassablePresets", "weakNegativePresets", "sourceRows", "duplicateCount",
        ])
        for p in payload["points"]:
            writer.writerow([
                p["id"], p["lat"], p["lng"],
                "|".join(map(str, p["observedTruckClasses"])),
                "|".join(p["observedPositivePresets"]),
                "|".join(p["inferredPassablePresets"]),
                "|".join(p["weakNegativePresets"]),
                "|".join(map(str, p["sourceRows"])),
                p["duplicateCount"],
            ])


def import_teacher_points(xlsx: Path, sheet: str | None) -> dict:
    sheet_name, rows = read_workbook_rows(xlsx, sheet)
    points, errors = build_points(rows)
    counts: dict[str, int] = defaultdict(int)
    for p in points:
        for truck_class in p["observedTruckClasses"]:
            counts[str(truck_class)] += 1
    return {
        "version": 1,
        "source": {
            "file": str(xlsx),
            "sheet": sheet_name,
            "validRows": len([r for r in rows if "error" not in r]),
            "invalidRows": len(errors),
            "generatedAt": datetime.now(timezone.utc).isoformat(),
        },
        "labelPolicy": {
            "observedPositivePresets": "Strong positive: this truck class was recorded as having entered.",
            "inferredPassablePresets": "Monotonic inference: classes up to the largest observed class are treated as passable candidates.",
            "weakNegativePresets": "Larger unrecorded classes are weak negatives only, not hard NG labels unless separately confirmed.",
        },
        "vehicleOrder": [PRESET_BY_CLASS[n] for n in TRUCK_CLASSES],
        "observedPointCountsByTruckClass": dict(sorted(counts.items(), key=lambda kv: int(kv[0]))),
        "errors": errors,
        "points": points,
    }


def selfcheck() -> int:
    try:
        import openpyxl
    except ImportError as exc:
        print(f"[FAIL] openpyxl import: {exc}")
        return 1
    with tempfile.TemporaryDirectory() as tmp:
        tmp_path = Path(tmp) / "teacher_selfcheck.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["トラック結果", "緯度", "経度"])
        ws.append([2, 35.0, 139.0])
        ws.append([4, 35.1, 139.1])
        ws.append(["4t", 35.1, 139.1])
        ws.append(["bad", 91, 139.1])
        wb.save(tmp_path)
        wb.close()
        payload = import_teacher_points(tmp_path, None)
        checks = [
            ("valid unique points", len(payload["points"]) == 2),
            ("duplicate merged", payload["points"][1]["duplicateCount"] == 2),
            ("4t inferred pass includes 2/3/4", payload["points"][1]["inferredPassablePresets"] == ["2t_flat", "3t_flat", "4t_flat"]),
            ("4t weak negative is 10t", payload["points"][1]["weakNegativePresets"] == ["10t_unic"]),
            ("invalid row captured", payload["source"]["invalidRows"] == 1),
        ]
        ok = True
        for name, passed in checks:
            print(f"[{'PASS' if passed else 'FAIL'}] {name}")
            ok = ok and passed
        print("\nselfcheck ALL PASS" if ok else "\nselfcheck FAILED")
        return 0 if ok else 1


def main(argv: list[str]) -> int:
    args = parse_args(argv)
    if args.selfcheck:
        return selfcheck()

    xlsx = Path(args.xlsx).resolve() if args.xlsx else find_default_xlsx()
    payload = import_teacher_points(xlsx, args.sheet)
    print(
        f"points={len(payload['points'])} validRows={payload['source']['validRows']} "
        f"invalidRows={payload['source']['invalidRows']} counts={payload['observedPointCountsByTruckClass']}"
    )
    if payload["errors"]:
        for err in payload["errors"][:10]:
            print(f"[skip] row {err.get('row')}: {err.get('error')}")
        if len(payload["errors"]) > 10:
            print(f"[skip] ... {len(payload['errors']) - 10} more")
    if args.dry_run:
        print("--dry-run: no files written")
        return 0
    write_outputs(payload, Path(args.out).resolve(), Path(args.csv_out).resolve())
    print(f"saved: {Path(args.out).resolve()}")
    print(f"saved: {Path(args.csv_out).resolve()}")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
